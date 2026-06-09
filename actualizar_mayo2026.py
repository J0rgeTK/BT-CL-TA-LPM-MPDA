"""Actualiza datos y salidas con afluencia real mayo 2026."""
from pathlib import Path
import re
import numpy as np
import pandas as pd
import openpyxl

import pipeline_afluencia as P
import oferta as O

ROOT = Path(__file__).resolve().parent
DATA = ROOT / 'data'
OUT = ROOT / 'outputs'
OUT.mkdir(exist_ok=True)

RAW = DATA / 'raw_mayo2026'
MAY_FILES = {
    'BIOTREN': RAW / '0. Afluencia diaria Biotren may_2026.xlsx',
    'CORTO_LAJA': RAW / 'Afluencia Laja Talcahuano 05-26.xlsx',
    'TREN_ARAUCANIA': RAW / '5. Reporte Tren Araucanía mayo 2026.xlsx',
    'LLANQUIHUE_PM': RAW / '5. Reporte Llanquihue - Puerto Montt mayo 2026.xlsx',
}


def parse_biotren_mayo(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = [s for s in wb.sheetnames if 'diaria' in s.lower()][0]
    rows = list(wb[sh].iter_rows(values_only=True))
    h = next(i for i, r in enumerate(rows) if r and len(r) > 1 and str(r[1]).strip().upper() == 'FECHA')
    hdr = rows[h]
    total_col = next(j for j, v in enumerate(hdr) if isinstance(v, str) and v.strip().lower() == 'afluencias + multas +sse')
    rec = []
    for r in rows[h + 1:]:
        if not r or len(r) < 2 or r[1] is None:
            continue
        fecha = pd.to_datetime(r[1], errors='coerce')
        if pd.isna(fecha) or fecha.year != 2026 or fecha.month != 5:
            continue
        rec.append(('BIOTREN', fecha.normalize(), P._num(r[total_col]) if total_col < len(r) else 0.0))
    return rec


def parse_matrix_mayo(path, sheet, servicio):
    rec = P._parse_matrix(path, sheet, servicio)
    df = pd.DataFrame(rec, columns=['servicio', 'fecha', 'pasajeros'])
    if df.empty:
        return df
    df['fecha'] = pd.to_datetime(df['fecha'])
    return df[(df['fecha'].dt.year == 2026) & (df['fecha'].dt.month == 5)]


def parse_matrix_detalle(path, sheet, servicio):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    hdr = rows[0]
    cols = [i for i, h in enumerate(hdr) if h is not None and re.match(r'^\d{4,6}$', str(h).strip())]
    out = []
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        fecha = pd.to_datetime(r[0], errors='coerce')
        if pd.isna(fecha) or fecha.year != 2026 or fecha.month != 5:
            continue
        vals = [P._num(r[i]) if i < len(r) else 0.0 for i in cols]
        out.append({
            'servicio': servicio,
            'fecha': fecha.normalize(),
            'pasajeros': sum(vals),
            'servicios_con_afluencia': sum(1 for v in vals if v > 0),
            'columnas_servicio': len(cols),
        })
    return pd.DataFrame(out)


def dt_label(fecha):
    dow = fecha.dayofweek
    return 'LV' if dow < 5 else ('Sab' if dow == 5 else 'Dom')


# 1) Consolidar mayo 2026.
may = pd.DataFrame(parse_biotren_mayo(MAY_FILES['BIOTREN']), columns=['servicio', 'fecha', 'pasajeros'])
may = pd.concat([
    may,
    parse_matrix_mayo(MAY_FILES['CORTO_LAJA'], 'Afxdia-may26', 'CORTO_LAJA'),
    parse_matrix_mayo(MAY_FILES['TREN_ARAUCANIA'], 'AfluDiariaxServ', 'TREN_ARAUCANIA'),
    parse_matrix_mayo(MAY_FILES['LLANQUIHUE_PM'], 'AfluDiariaxServ', 'LLANQUIHUE_PM'),
], ignore_index=True)
may = may[may['pasajeros'] > 0].copy()
may['fecha'] = pd.to_datetime(may['fecha'])
may['dt'] = may['fecha'].apply(dt_label)
may.to_csv(DATA / 'afluencia_mayo_2026_cargada.csv', index=False)

old = pd.read_csv(DATA / 'afluencia_diaria_consolidada.csv', parse_dates=['fecha'])
new = (pd.concat([old, may[['servicio', 'fecha', 'pasajeros']]], ignore_index=True)
       .sort_values('pasajeros')
       .drop_duplicates(['servicio', 'fecha'], keep='last')
       .sort_values(['servicio', 'fecha'])
       .reset_index(drop=True))
new.to_csv(DATA / 'afluencia_diaria_consolidada.csv', index=False)

# 2) Resumen de mayo y servicios observados por matrices.
matrices = pd.concat([
    parse_matrix_detalle(MAY_FILES['CORTO_LAJA'], 'Afxdia-may26', 'CORTO_LAJA'),
    parse_matrix_detalle(MAY_FILES['TREN_ARAUCANIA'], 'AfluDiariaxServ', 'TREN_ARAUCANIA'),
    parse_matrix_detalle(MAY_FILES['LLANQUIHUE_PM'], 'AfluDiariaxServ', 'LLANQUIHUE_PM'),
], ignore_index=True)
matrices['dt'] = matrices['fecha'].apply(dt_label)

nd_2026 = O.dias_por_tipo(2026)
may_days = nd_2026[nd_2026['mes'] == 5].set_index('dt')['n_dias'].to_dict()
# Oferta programada vigente usada por modelo para mayo 2026.
servicios_mayo_modelo = {
    'BIOTREN': may_days['LV'] * (40 + 106) + may_days['Sab'] * (8 + 53) + may_days['Dom'] * (0 + 32),
    'CORTO_LAJA': may_days['LV'] * 10 + may_days['Sab'] * 8 + may_days['Dom'] * 8,
    'TREN_ARAUCANIA': may_days['LV'] * 18.6 + may_days['Sab'] * 12 + may_days['Dom'] * 6,
    'LLANQUIHUE_PM': may_days['LV'] * 20,
}
servicios_obs = matrices.groupby('servicio')['servicios_con_afluencia'].sum().to_dict()
resumen = may.groupby('servicio').agg(dias_obs=('fecha', 'nunique'), afluencia_mayo_2026=('pasajeros', 'sum')).reset_index()
resumen['servicios_programados_modelo_mayo'] = resumen['servicio'].map(servicios_mayo_modelo)
resumen['servicios_con_afluencia_matriz'] = resumen['servicio'].map(servicios_obs)
resumen['pax_viaje_modelo_mayo'] = resumen['afluencia_mayo_2026'] / resumen['servicios_programados_modelo_mayo']
resumen['pax_viaje_matriz_mayo'] = resumen['afluencia_mayo_2026'] / resumen['servicios_con_afluencia_matriz']
resumen.to_csv(DATA / 'resumen_mayo_2026.csv', index=False)

# 3) Calibracion de productividad pax/viaje por tipo de dia.
params_base = O.aplicar_oferta_actual(pd.read_csv(DATA / 'oferta_params.csv'))

def pxv_biotren_modelo(dt):
    sub = params_base[(params_base['unit'].isin(['BIOTREN_L1', 'BIOTREN_L2'])) & (params_base['dt'] == dt)]
    return (sub['servicios_dia'] * sub['pax_x_viaje']).sum() / sub['servicios_dia'].sum()

# Observado mayo por tipo de dia.
bt_dt = may[may['servicio'] == 'BIOTREN'].groupby('dt')['pasajeros'].sum().to_dict()
cal_rows = []
servicios_bt_dt = {'LV': 146, 'Sab': 61, 'Dom': 32}
for dt, servicios_dia in servicios_bt_dt.items():
    obs_pxv = bt_dt.get(dt, 0) / max(servicios_dia * may_days[dt], 1)
    base_pxv = pxv_biotren_modelo(dt)
    factor = obs_pxv / base_pxv if base_pxv else 1.0
    for unit in ['BIOTREN_L1', 'BIOTREN_L2']:
        cal_rows.append({'unit': unit, 'servicio': 'BIOTREN', 'dt': dt,
                         'pax_viaje_observado_mayo': obs_pxv,
                         'pax_viaje_modelo_base': base_pxv,
                         'factor_objetivo': factor,
                         'peso_calibracion': 0.70,
                         'factor_min': 0.75,
                         'factor_max': 1.25,
                         'criterio': 'Factor agregado Biotren aplicado a L1 y L2 por no disponer de afluencia oficial separada por linea.'})

# Servicios con matriz por tren.
for unit, peso in [('CORTO_LAJA', 0.60), ('TREN_ARAUCANIA', 0.70), ('LLANQUIHUE_PM', 0.80)]:
    d = matrices[matrices['servicio'] == unit]
    for dt, g in d.groupby('dt'):
        if unit == 'LLANQUIHUE_PM' and dt != 'LV':
            continue
        # Para servicios regionales se usa conteo de servicios de la matriz cuando existe;
        # para Tren Araucania se mantiene la oferta promedio del modelo por tipo de dia,
        # pues el objetivo es comparar contra el itinerario vigente informado.
        if unit == 'TREN_ARAUCANIA':
            servicios = {'LV': 18.6, 'Sab': 12, 'Dom': 6}[dt] * may_days[dt]
        elif unit == 'CORTO_LAJA':
            servicios = {'LV': 10, 'Sab': 8, 'Dom': 8}[dt] * may_days[dt]
        else:
            servicios = g['servicios_con_afluencia'].sum()
        obs_pxv = g['pasajeros'].sum() / max(servicios, 1)
        base_pxv = params_base[(params_base['unit'] == unit) & (params_base['dt'] == dt)]['pax_x_viaje'].mean()
        factor = obs_pxv / base_pxv if base_pxv else 1.0
        cal_rows.append({'unit': unit, 'servicio': unit, 'dt': dt,
                         'pax_viaje_observado_mayo': obs_pxv,
                         'pax_viaje_modelo_base': base_pxv,
                         'factor_objetivo': factor,
                         'peso_calibracion': peso,
                         'factor_min': 0.65,
                         'factor_max': 1.45,
                         'criterio': 'Calibracion parcial con mayo 2026 por tipo de dia.'})
cal = pd.DataFrame(cal_rows)
cal['factor_aplicado'] = np.minimum(cal['factor_max'], np.maximum(cal['factor_min'], 1 + cal['peso_calibracion'] * (cal['factor_objetivo'] - 1)))
cal.to_csv(DATA / 'calibracion_mayo_2026.csv', index=False)

# 4) Mensualizar y proyectar.
mdf = P.mensualizar(new)
mdf.to_csv(DATA / 'afluencia_mensual_modelo.csv', index=False)
base, meta = P.proyectar_2027(mdf)
base.to_csv(OUT / 'proyeccion_2027_referencia_estacional.csv')

params = O.aplicar_oferta_actual(pd.read_csv(DATA / 'oferta_params.csv'))
uni, serv = O.proyectar_conservador(params, base_servicios=base)
_, solo_oferta = O.proyectar(params)
# Nombres vigentes del escenario recomendado. Se mantienen tambien los archivos
# historicos con sufijo conservador por compatibilidad con versiones previas.
uni.to_csv(OUT / 'proyeccion_2027_unidades_base_ajustada.csv')
serv.to_csv(OUT / 'proyeccion_2027_resumen_base_ajustada.csv')
uni.to_csv(OUT / 'proyeccion_2027_unidades_conservador.csv')
serv.to_csv(OUT / 'proyeccion_2027_resumen_conservador.csv')
comp = pd.DataFrame({
    'ref_estacional': base.sum(),
    'solo_oferta_calibrada': solo_oferta.sum(),
    'base_calibrada_ajustada': serv.sum(),
})
comp['dif_vs_oferta'] = comp['base_calibrada_ajustada'] - comp['solo_oferta_calibrada']
comp['dif_vs_ref'] = comp['base_calibrada_ajustada'] - comp['ref_estacional']
comp.to_csv(OUT / 'comparativo_escenarios_2027.csv')

validacion = resumen.merge(comp, left_on='servicio', right_index=True, how='left')
validacion.to_csv(OUT / 'validacion_mayo_2026_vs_proyeccion.csv', index=False)

print('Mayo 2026 cargado:')
print(resumen.to_string(index=False))
print('\nComparativo 2027:')
print(comp.to_string())
